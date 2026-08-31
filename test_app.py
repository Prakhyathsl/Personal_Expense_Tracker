"""Integration tests isolated to data_test; never touches production data."""
import os
from pathlib import Path
import shutil, zipfile, io

BASE_DIR = Path(__file__).resolve().parent
TEST_DATA = BASE_DIR / 'data_test'
if TEST_DATA.exists():
    shutil.rmtree(TEST_DATA)
TEST_DATA.mkdir(parents=True)

# app.py reads this before importing/initializing its database.
os.environ['APP_DATA_DIR'] = str(TEST_DATA)
os.environ.pop('DATABASE_URL', None)
os.environ['SECRET_KEY'] = 'test-secret'

from openpyxl import load_workbook
import app as app_module
import auth_db, db

app_module.DATA_DIR = TEST_DATA
app_module.AUTH_SQLITE = TEST_DATA / 'auth.db'
auth_db.AUTH_SQLITE = TEST_DATA / 'auth.db'
db.DATA_DIR = TEST_DATA
db.AUTH_SQLITE = TEST_DATA / 'auth.db'
app_module.app.config.update(TESTING=True, SECRET_KEY='test-secret')
app_module.init_database()
client = app_module.app.test_client()

def check(r, code):
    assert r.status_code==code, f'Expected {code}, got {r.status_code}: {r.data!r}'

def main():
    # 1-2 protected access
    check(client.get('/'),302); assert '/login' in client.get('/').location
    check(client.get('/api/expenses'),401)
    print('PASS: unauthenticated dashboard/API protection')
    # 3 registration + 31 first admin
    r=client.post('/api/auth/register',json={'name':'Test Student','username':'teststudent','email':'test@example.com','password':'Password123'}); check(r,201)
    u=r.get_json()['user']; assert u['is_admin'] is True
    # 4 duplicate + case-insensitive
    r=client.post('/api/auth/register',json={'name':'Other','username':'TESTSTUDENT','email':'other@example.com','password':'Password123'}); assert r.status_code in (400,409)
    r=client.post('/api/auth/register',json={'name':'Other','username':'otheruser','email':'TEST@EXAMPLE.COM','password':'Password123'}); assert r.status_code in (400,409)
    print('PASS: registration, first-admin, duplicate prevention')
    # 5-8 empty/dashboard/stats
    check(client.get('/'),200); html=client.get('/').get_data(as_text=True); assert 'class="page active" id="page-dashboard"' in html
    check(client.get('/api/expenses'),200); assert client.get('/api/expenses').get_json()==[]
    s=client.get('/api/stats').get_json(); assert float(s['total_spending'])==0 and float(s['total_expenses'])==0 and s['transaction_count']==0
    assert s['average_expense']==0 and s['highest_expense']==0 and s['monthly_average']==0
    print('PASS: fresh account starts with zero financial data')
    # 9-10 add/read
    expense={'academic_year':'1st Year','semester':'1st Sem','date':'30-08-2026','category':'Food','amount':120.50,'payment_method':'UPI','description':'Lunch'}
    r=client.post('/api/expenses',json=expense); check(r,201); eid=r.get_json()['expense']['id']; assert len(client.get('/api/expenses').get_json())==1
    # 11-12 update/stats
    updated={**expense,'amount':200,'description':'Dinner'}; check(client.put(f'/api/expenses/{eid}',json=updated),200); xs=client.get('/api/expenses').get_json(); assert len(xs)==1; assert float(xs[0]['amount'])==200; assert abs(client.get('/api/stats').get_json()['total_spending']-200)<1e-6
    # 13 delete and verify stats return to zero
    check(client.delete(f'/api/expenses/{eid}'),200); assert client.get('/api/expenses').get_json()==[]; s=client.get('/api/stats').get_json(); assert s['transaction_count']==0 and float(s['total_spending'])==0
    print('PASS: expense CRUD and stats update')
    # 14-16 category/academic/semester endpoints
    cats=client.get('/api/categories'); check(cats,200); assert 'Food' in cats.get_json()['categories']
    ay=client.get('/api/academic-years'); check(ay,200); assert ay.get_json()['year_semesters']['1st Year']==['1st Sem','2nd Sem']
    assert '8th Sem' in ay.get_json()['academic_years'] or '8th Sem' in sum(ay.get_json()['year_semesters'].values(),[])
    print('PASS: categories, academic years and semesters')
    # 17-19 goal/contribution/progress
    r=client.post('/api/savings/goals',json={'name':'New Phone','target_amount':1000,'target_date':'30-12-2026','description':'Goal'}); check(r,201); gid=r.get_json()['goal']['id']
    r=client.post('/api/savings',json={'goal_id':gid,'amount':250,'date':'30-08-2026','note':'Monthly saving'}); check(r,201)
    sv=client.get('/api/savings').get_json(); assert len(sv['history'])==1; assert abs(sv['goals'][0]['progress']-25)<1e-6
    # goal edit
    check(client.put(f'/api/savings/goals/{gid}',json={'name':'Phone','target_date':'31-12-2026'}),200)
    print('PASS: savings goal, contribution, progress and edit')
    # 20 delete saving; 21 goal
    sid=sv['history'][0]['id']; check(client.delete(f'/api/savings/{sid}'),200); check(client.delete(f'/api/savings/goals/{gid}'),200)
    # 22 backup, 23 excel
    # Add exactly one expense so export contains header + one data row.
    r=client.post('/api/expenses',json=expense); check(r,201); assert len(client.get('/api/expenses').get_json())==1
    r=client.get('/api/backup'); check(r,200); assert r.mimetype=='application/zip'; z=zipfile.ZipFile(io.BytesIO(r.data)); assert set(z.namelist())=={'Expense_Tracker_Master.xlsx','Savings_Tracker.xlsx'}; z.close()
    r=client.get('/api/export/expenses'); check(r,200); assert r.mimetype=='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'; wb=load_workbook(io.BytesIO(r.data),read_only=True); ws=wb['Expenses']; assert ws.max_row==2; assert [c.value for c in ws[1]]==app_module.HEADERS; assert ws['D2'].value.strftime('%d-%m-%Y')=='30-08-2026'; assert float(ws['F2'].value)==120.50; wb.close()
    print('PASS: backup and Excel export')
    # 24 forgot password no SMTP
    r=client.post('/api/auth/forgot-password',json={'email':'test@example.com'}); check(r,200); assert 'success' in r.get_json()
    # 25 logout, 26 protected
    check(client.post('/api/auth/logout'),200); check(client.get('/'),302); check(client.get('/api/expenses'),401)
    print('PASS: logout and protected endpoints after logout')
    # 27 wrong password, 28 correct login
    r=client.post('/api/auth/login',json={'identifier':'teststudent','password':'wrong'}); check(r,401)
    r=client.post('/api/auth/login',json={'identifier':'TESTSTUDENT','password':'Password123'}); check(r,200); assert r.get_json()['user']['is_admin'] is True
    print('PASS: wrong password rejected and case-insensitive login works')
    # 29-33 second user isolation/admin/date validation/fresh data
    r=client.post('/api/auth/register',json={'name':'Second Student','username':'seconduser','email':'second@example.com','password':'Password123'}); check(r,201); u2=r.get_json()['user']; assert u2['is_admin'] is False
    assert client.get('/api/expenses').get_json()==[]
    bad={**expense,'date':'1212222'}; r=client.post('/api/expenses',json=bad); assert r.status_code==400
    bad2={**expense,'amount':-1}; assert client.post('/api/expenses',json=bad2).status_code==400
    # The live version stores financial data in the database; exports are generated on demand.
    r=client.get('/api/export/expenses'); check(r,200); wb=load_workbook(io.BytesIO(r.data),read_only=True); assert wb['Expenses'].max_row==1; wb.close()
    # logout and login first user; first user has its own expense
    client.post('/api/auth/logout'); check(client.post('/api/auth/login',json={'identifier':'test@example.com','password':'Password123'}),200); assert len(client.get('/api/expenses').get_json())==1
    print('PASS: per-user isolation, date validation, second user non-admin and fresh data')
    print('\n'+'='*60+'\nALL PERSONAL EXPENSE TRACKER TESTS PASSED\n'+'='*60)

try:
    main()
finally:
    shutil.rmtree(TEST_DATA,ignore_errors=True)
    os.environ.pop('APP_DATA_DIR', None)
