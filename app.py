import os, re, tempfile, zipfile, io
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from functools import wraps
from threading import RLock

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, after_this_request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import db
from db import connection, ph, using_postgres
from auth_db import init_auth_db, register_user, authenticate_user, get_user_by_id, request_password_reset, reset_password

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
AUTH_SQLITE = DATA_DIR / 'auth.db'
SHEET_NAME = 'Expenses'
GOALS_SHEET = 'Savings Goals'
HISTORY_SHEET = 'Savings History'
HEADERS = ['ID','Academic Year','Semester','Date','Category','Amount','Payment Method','Description']
GOAL_HEADERS = ['ID','Goal Name','Target Amount','Target Date','Description','Created Date']
HISTORY_HEADERS = ['ID','Goal ID','Amount','Date','Note']
CATEGORIES = ['Food','Transport','Education','Shopping','Entertainment','Health','Bills','Accommodation','Other']
PAYMENT_METHODS = ['UPI','Cash','Card','Net Banking','Other']
ACADEMIC_YEARS = ['1st Year','2nd Year','3rd Year','4th Year']
SEMESTERS = ['1st Sem','2nd Sem','3rd Sem','4th Sem','5th Sem','6th Sem','7th Sem','8th Sem']
YEAR_SEMESTERS = {'1st Year':['1st Sem','2nd Sem'],'2nd Year':['3rd Sem','4th Sem'],'3rd Year':['5th Sem','6th Sem'],'4th Year':['7th Sem','8th Sem']}
FILE_LOCK = RLock()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-in-production')
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.environ.get('COOKIE_SECURE', '0') == '1'
)


def init_database():
    init_auth_db()
    with connection(sqlite_path=DATA_DIR / 'auth.db') as conn:
        cur = conn.cursor(); p = ph()
        if using_postgres():
            cur.execute('''CREATE TABLE IF NOT EXISTS expenses (
                id BIGSERIAL PRIMARY KEY, user_id BIGINT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                academic_year TEXT NOT NULL, semester TEXT NOT NULL, date TEXT NOT NULL,
                category TEXT NOT NULL, amount NUMERIC(14,2) NOT NULL CHECK (amount > 0),
                payment_method TEXT NOT NULL, description TEXT NOT NULL DEFAULT '', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS savings_goals (
                id BIGSERIAL PRIMARY KEY, user_id BIGINT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                name TEXT NOT NULL, target_amount NUMERIC(14,2) NOT NULL CHECK (target_amount > 0),
                target_date TEXT NOT NULL, description TEXT NOT NULL DEFAULT '', created_date TEXT NOT NULL
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS savings_history (
                id BIGSERIAL PRIMARY KEY, user_id BIGINT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                goal_id BIGINT NOT NULL REFERENCES savings_goals(id) ON DELETE CASCADE,
                amount NUMERIC(14,2) NOT NULL CHECK (amount > 0), date TEXT NOT NULL, note TEXT NOT NULL DEFAULT ''
            )''')
        else:
            cur.execute('''CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY, user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                academic_year TEXT NOT NULL, semester TEXT NOT NULL, date TEXT NOT NULL,
                category TEXT NOT NULL, amount REAL NOT NULL CHECK (amount > 0),
                payment_method TEXT NOT NULL, description TEXT NOT NULL DEFAULT '', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS savings_goals (
                id INTEGER PRIMARY KEY, user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                name TEXT NOT NULL, target_amount REAL NOT NULL CHECK (target_amount > 0),
                target_date TEXT NOT NULL, description TEXT NOT NULL DEFAULT '', created_date TEXT NOT NULL
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS savings_history (
                id INTEGER PRIMARY KEY, user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                goal_id INTEGER NOT NULL REFERENCES savings_goals(id) ON DELETE CASCADE,
                amount REAL NOT NULL CHECK (amount > 0), date TEXT NOT NULL, note TEXT NOT NULL DEFAULT ''
            )''')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_expenses_user ON expenses(user_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_goals_user ON savings_goals(user_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_history_user ON savings_history(user_id)')


# Keep initialization automatic for both local development and Render.
init_database()


def current_user():
    uid = session.get('user_id')
    return get_user_by_id(uid) if uid else None


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user():
            if request.path.startswith('/api/'):
                return jsonify({'error':'Authentication required.','authenticated':False}), 401
            return redirect(url_for('login_page'))
        return view(*args, **kwargs)
    return wrapped


def user_id():
    u = current_user()
    if not u:
        raise RuntimeError('Authentication required')
    return int(u['id'])


def parse_date(value, field='Date'):
    value = str(value or '').strip()
    if not re.fullmatch(r'\d{2}-\d{2}-\d{4}', value):
        raise ValueError(f'{field} must be a valid date in DD-MM-YYYY format.')
    try:
        return datetime.strptime(value, '%d-%m-%Y').strftime('%d-%m-%Y')
    except ValueError:
        raise ValueError(f'{field} must be a valid date in DD-MM-YYYY format.')


def validate_expense(data):
    if not isinstance(data, dict): raise ValueError('Invalid request data.')
    ay = str(data.get('academic_year','')).strip(); sem = str(data.get('semester','')).strip()
    cat = str(data.get('category','')).strip(); pm = str(data.get('payment_method','')).strip(); desc = str(data.get('description','')).strip()
    if ay not in ACADEMIC_YEARS: raise ValueError('Please select a valid academic year.')
    if sem not in YEAR_SEMESTERS[ay]: raise ValueError('Please select a valid semester for the selected academic year.')
    if cat not in CATEGORIES: raise ValueError('Please select a valid category.')
    if pm not in PAYMENT_METHODS: raise ValueError('Please select a valid payment method.')
    date = parse_date(data.get('date'))
    try: amount = float(data.get('amount'))
    except (ValueError, TypeError): raise ValueError('Amount must be a positive number.')
    if amount <= 0: raise ValueError('Amount must be a positive number.')
    if amount > 999999999: raise ValueError('Amount is too large.')
    return {'academic_year':ay,'semester':sem,'date':date,'category':cat,'amount':round(amount,2),'payment_method':pm,'description':desc}


def row_dict(cur, row):
    if row is None: return None
    if using_postgres(): return dict(zip([d[0] for d in cur.description], row))
    return dict(row)


def expense_from_row(cur, row):
    x = row_dict(cur, row)
    return {'id':int(x['id']), 'academic_year':x['academic_year'], 'semester':x['semester'], 'date':x['date'], 'category':x['category'], 'amount':float(x['amount']), 'payment_method':x['payment_method'], 'description':x['description']}


def get_expenses():
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph(); cur.execute(f'''SELECT id,academic_year,semester,date,category,amount,payment_method,description
            FROM expenses WHERE user_id={p} ORDER BY id''',(user_id(),))
        return [expense_from_row(cur,r) for r in cur.fetchall()]


def next_local_id(cur, table):
    cur.execute(f'SELECT COALESCE(MAX(id),0)+1 FROM {table}')
    return int(cur.fetchone()[0])


def create_expense_db(item):
    uid=user_id()
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph()
        if using_postgres():
            cur.execute('''INSERT INTO expenses (user_id,academic_year,semester,date,category,amount,payment_method,description)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                        (uid,item['academic_year'],item['semester'],item['date'],item['category'],item['amount'],item['payment_method'],item['description']))
            item['id']=int(cur.fetchone()[0])
        else:
            eid=next_local_id(cur,'expenses'); item['id']=eid
            cur.execute(f'''INSERT INTO expenses (id,user_id,academic_year,semester,date,category,amount,payment_method,description)
                            VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p})''',
                        (eid,uid,item['academic_year'],item['semester'],item['date'],item['category'],item['amount'],item['payment_method'],item['description']))
    return item


def update_expense_db(eid,item):
    uid=user_id()
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph()
        cur.execute(f'''UPDATE expenses SET academic_year={p},semester={p},date={p},category={p},amount={p},payment_method={p},description={p}
                        WHERE id={p} AND user_id={p}''',
                    (item['academic_year'],item['semester'],item['date'],item['category'],item['amount'],item['payment_method'],item['description'],eid,uid))
        if cur.rowcount != 1: return False
    item['id']=eid; return True


def delete_expense_db(eid):
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph(); cur.execute(f'DELETE FROM expenses WHERE id={p} AND user_id={p}',(eid,user_id())); return cur.rowcount==1


def savings_data():
    uid=user_id()
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph()
        cur.execute(f'''SELECT id,name,target_amount,target_date,description,created_date FROM savings_goals WHERE user_id={p} ORDER BY id''',(uid,))
        goals=[]
        for r in cur.fetchall():
            x=row_dict(cur,r); goals.append({'id':int(x['id']),'name':x['name'],'target_amount':float(x['target_amount']),'target_date':x['target_date'],'description':x['description'],'created_date':x['created_date']})
        cur.execute(f'''SELECT id,goal_id,amount,date,note FROM savings_history WHERE user_id={p} ORDER BY id''',(uid,))
        history=[]
        for r in cur.fetchall():
            x=row_dict(cur,r); history.append({'id':int(x['id']),'goal_id':int(x['goal_id']),'amount':float(x['amount']),'date':x['date'],'note':x['note']})
    for g in goals:
        saved=round(sum(x['amount'] for x in history if x['goal_id']==g['id']),2)
        g['contributed']=saved; g['remaining']=round(max(g['target_amount']-saved,0),2); g['progress']=round(min(saved/g['target_amount']*100,100),2) if g['target_amount'] else 0
    return goals,history


def create_goal_db(name,target,target_date,desc):
    uid=user_id(); created=datetime.now().strftime('%d-%m-%Y')
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph()
        if using_postgres():
            cur.execute('''INSERT INTO savings_goals (user_id,name,target_amount,target_date,description,created_date)
                           VALUES (%s,%s,%s,%s,%s,%s) RETURNING id''',(uid,name,target,target_date,desc,created)); gid=int(cur.fetchone()[0])
        else:
            gid=next_local_id(cur,'savings_goals'); cur.execute(f'''INSERT INTO savings_goals (id,user_id,name,target_amount,target_date,description,created_date)
                VALUES ({p},{p},{p},{p},{p},{p},{p})''',(gid,uid,name,target,target_date,desc,created))
    return {'id':gid,'name':name,'target_amount':target,'target_date':target_date,'description':desc,'created_date':created,'contributed':0,'remaining':target,'progress':0}


def add_saving_db(gid,amount,date,note):
    uid=user_id(); goals,history=savings_data(); goal=next((g for g in goals if g['id']==gid),None)
    if not goal: return None,'Savings goal not found.'
    remaining=max(goal['target_amount']-sum(x['amount'] for x in history if x['goal_id']==gid),0)
    if amount>remaining: return None,'This amount is greater than the remaining goal amount.'
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor(); p=ph()
        if using_postgres():
            cur.execute('''INSERT INTO savings_history (user_id,goal_id,amount,date,note) VALUES (%s,%s,%s,%s,%s) RETURNING id''',(uid,gid,amount,date,note)); sid=int(cur.fetchone()[0])
        else:
            sid=next_local_id(cur,'savings_history'); cur.execute(f'''INSERT INTO savings_history (id,user_id,goal_id,amount,date,note) VALUES ({p},{p},{p},{p},{p},{p})''',(sid,uid,gid,amount,date,note))
    return {'id':sid,'goal_id':gid,'amount':amount,'date':date,'note':note},None


def excel_expenses(items):
    wb=Workbook(); ws=wb.active; ws.title=SHEET_NAME; ws.append(HEADERS)
    for x in items:
        ws.append([x['id'],x['academic_year'],x['semester'],datetime.strptime(x['date'],'%d-%m-%Y'),x['category'],x['amount'],x['payment_method'],x['description']])
    format_ws(ws); return wb


def excel_savings(goals,history):
    wb=Workbook(); g=wb.active; g.title=GOALS_SHEET; g.append(GOAL_HEADERS)
    for x in goals: g.append([x['id'],x['name'],x['target_amount'],datetime.strptime(x['target_date'],'%d-%m-%Y'),x['description'],datetime.strptime(x['created_date'],'%d-%m-%Y')])
    h=wb.create_sheet(HISTORY_SHEET); h.append(HISTORY_HEADERS)
    for x in history: h.append([x['id'],x['goal_id'],x['amount'],datetime.strptime(x['date'],'%d-%m-%Y'),x['note']])
    format_ws(g); format_ws(h); return wb


def format_ws(ws):
    for c in ws[1]: c.font=Font(bold=True); c.fill=PatternFill('solid',fgColor='DDEBFF'); c.alignment=Alignment(horizontal='center')
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions


def workbook_bytes(wb):
    bio=io.BytesIO(); wb.save(bio); wb.close(); bio.seek(0); return bio


def json_error(msg,status=400): return jsonify({'error':msg}),status

@app.route('/')
@login_required
def home(): return render_template('index.html',categories=CATEGORIES,payment_methods=PAYMENT_METHODS,academic_years=ACADEMIC_YEARS,semesters=SEMESTERS,year_semesters=YEAR_SEMESTERS,current_user=current_user())
@app.route('/login')
def login_page(): return redirect(url_for('home')) if current_user() else render_template('login.html')
@app.route('/register')
def register_page(): return redirect(url_for('home')) if current_user() else render_template('register.html')
@app.route('/forgot-password')
def forgot_password_page(): return render_template('forgot_password.html')
@app.route('/reset-password/<token>')
def reset_password_page(token): return render_template('reset_password.html',token=token)

@app.route('/api/auth/register',methods=['POST'])
def api_register():
    try:
        d=request.get_json(silent=True) or {}; ok,result=register_user(d.get('name'),d.get('username'),d.get('email'),d.get('password'))
        if not ok:return json_error(result,409 if 'already registered' in result else 400)
        session.clear(); session['user_id']=result['id']; return jsonify({'success':True,'user':result}),201
    except Exception as e: return json_error(f'Unable to register: {e}',500)

@app.route('/api/auth/login',methods=['POST'])
def api_login():
    try:
        d=request.get_json(silent=True) or {}; u=authenticate_user(d.get('identifier'),d.get('password'))
        if not u:return json_error('Invalid username/email or password.',401)
        session.clear(); session['user_id']=u['id']; return jsonify({'success':True,'user':u})
    except Exception as e:return json_error(f'Unable to login: {e}',500)

@app.route('/api/auth/me')
def api_me():
    u=current_user(); return jsonify({'authenticated':True,'user':u}) if u else (jsonify({'authenticated':False}),401)
@app.route('/api/auth/logout',methods=['POST'])
def api_logout(): session.clear(); return jsonify({'success':True})
@app.route('/api/auth/forgot-password',methods=['POST'])
def api_forgot():
    d=request.get_json(silent=True) or {}; ok,msg=request_password_reset(d.get('email','')); return jsonify({'success':ok,'message':msg})
@app.route('/api/auth/reset-password',methods=['POST'])
def api_reset():
    d=request.get_json(silent=True) or {}; ok,msg=reset_password(d.get('token'),d.get('password')); return jsonify({'success':True,'message':msg}) if ok else json_error(msg)

@app.route('/api/expenses',methods=['GET'])
@login_required
def api_expenses():
    try:return jsonify(get_expenses())
    except Exception as e:return json_error(f'Unable to read expenses: {e}',500)
@app.route('/api/expenses',methods=['POST'])
@login_required
def api_add():
    try:return jsonify({'success':True,'expense':create_expense_db(validate_expense(request.get_json(silent=True))) }),201
    except ValueError as e:return json_error(str(e))
    except Exception as e:return json_error(f'Unable to save expense: {e}',500)
@app.route('/api/expenses/<int:eid>',methods=['PUT'])
@login_required
def api_update(eid):
    try:item=validate_expense(request.get_json(silent=True))
    except ValueError as e:return json_error(str(e))
    try:
        if not update_expense_db(eid,item):return json_error('Expense not found.',404)
        return jsonify({'success':True,'expense':item})
    except Exception as e:return json_error(f'Unable to update expense: {e}',500)
@app.route('/api/expenses/<int:eid>',methods=['DELETE'])
@login_required
def api_delete(eid):
    try:
        if not delete_expense_db(eid):return json_error('Expense not found.',404)
        return jsonify({'success':True})
    except Exception as e:return json_error(f'Unable to delete expense: {e}',500)

@app.route('/api/stats')
@login_required
def api_stats():
    items=get_expenses(); total=round(sum(x['amount'] for x in items),2); n=len(items); monthly=defaultdict(float)
    category=defaultdict(float); payment=defaultdict(float); years=defaultdict(float); sems=defaultdict(float)
    for x in items:
        dt=datetime.strptime(x['date'],'%d-%m-%Y'); monthly[dt.strftime('%Y-%m')]+=x['amount']; category[x['category']]+=x['amount']; payment[x['payment_method']]+=x['amount']; years[x['academic_year']]+=x['amount']; sems[x['semester']]+=x['amount']
    recent=sorted(items,key=lambda x:datetime.strptime(x['date'],'%d-%m-%Y'),reverse=True)[:5]
    return jsonify({'total_spending':total,'total_expenses':total,'transaction_count':n,'average_expense':round(total/n,2) if n else 0,'highest_expense':max([x['amount'] for x in items],default=0),'monthly_average':round(sum(monthly.values())/len(monthly),2) if monthly else 0,'category_totals':dict(category),'payment_totals':dict(payment),'year_totals':dict(years),'semester_totals':dict(sems),'recent_expenses':recent,'expenses':items})

@app.route('/api/categories')
@login_required
def api_categories(): return jsonify({'categories':CATEGORIES})
@app.route('/api/academic-years')
@login_required
def api_years(): return jsonify({'academic_years':ACADEMIC_YEARS,'year_semesters':YEAR_SEMESTERS})

@app.route('/api/savings',methods=['GET'])
@login_required
def api_savings_get():
    goals,history=savings_data(); return jsonify({'goals':goals,'history':history})
@app.route('/api/savings/goals',methods=['POST'])
@login_required
def api_goal_add():
    d=request.get_json(silent=True) or {}; name=str(d.get('name','')).strip(); desc=str(d.get('description','')).strip()
    if len(name)<1:return json_error('Goal name is required.')
    try:target=float(d.get('target_amount'))
    except (ValueError,TypeError):return json_error('Target amount must be a positive number.')
    if target<=0:return json_error('Target amount must be a positive number.')
    try:td=parse_date(d.get('target_date'),'Target date')
    except ValueError as e:return json_error(str(e))
    try:return jsonify({'success':True,'goal':create_goal_db(name,round(target,2),td,desc)}),201
    except Exception as e:return json_error(f'Unable to create savings goal: {e}',500)
@app.route('/api/savings/goals/<int:gid>',methods=['PUT'])
@login_required
def api_goal_update(gid):
    d=request.get_json(silent=True) or {}; goals,history=savings_data(); goal=next((x for x in goals if x['id']==gid),None)
    if not goal:return json_error('Savings goal not found.',404)
    if d.get('name') is not None:
        goal['name']=str(d['name']).strip()
        if not goal['name']:return json_error('Goal name is required.')
    if d.get('description') is not None:goal['description']=str(d['description']).strip()
    if d.get('target_amount') is not None:
        try:v=float(d['target_amount'])
        except (ValueError,TypeError):return json_error('Target amount must be positive.')
        if v<=0:return json_error('Target amount must be positive.')
        goal['target_amount']=round(v,2)
    if d.get('target_date') is not None:
        try:goal['target_date']=parse_date(d['target_date'],'Target date')
        except ValueError as e:return json_error(str(e))
    try:
        with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
            cur=conn.cursor();p=ph();cur.execute(f'''UPDATE savings_goals SET name={p},target_amount={p},target_date={p},description={p} WHERE id={p} AND user_id={p}''',(goal['name'],goal['target_amount'],goal['target_date'],goal['description'],gid,user_id()))
            if cur.rowcount!=1:return json_error('Savings goal not found.',404)
        goals,_=savings_data();return jsonify({'success':True,'goal':next(x for x in goals if x['id']==gid)})
    except Exception as e:return json_error(f'Unable to update savings goal: {e}',500)
@app.route('/api/savings',methods=['POST'])
@login_required
def api_save_add():
    d=request.get_json(silent=True) or {}
    try:gid=int(d.get('goal_id'));amount=float(d.get('amount'));date=parse_date(d.get('date'))
    except (ValueError,TypeError):return json_error('Invalid saving contribution data.')
    if amount<=0:return json_error('Saving amount must be positive.')
    item,err=add_saving_db(gid,round(amount,2),date,str(d.get('note','')).strip())
    if err:return json_error(err,404 if 'not found' in err.lower() else 400)
    return jsonify({'success':True,'savings':item}),201
@app.route('/api/savings/<int:sid>',methods=['DELETE'])
@login_required
def api_save_delete(sid):
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor();p=ph();cur.execute(f'DELETE FROM savings_history WHERE id={p} AND user_id={p}',(sid,user_id()))
        if cur.rowcount!=1:return json_error('Savings record not found.',404)
    return jsonify({'success':True})
@app.route('/api/savings/goals/<int:gid>',methods=['DELETE'])
@login_required
def api_goal_delete(gid):
    with connection(sqlite_path=DATA_DIR/'auth.db') as conn:
        cur=conn.cursor();p=ph();cur.execute(f'DELETE FROM savings_goals WHERE id={p} AND user_id={p}',(gid,user_id()))
        if cur.rowcount!=1:return json_error('Savings goal not found.',404)
    return jsonify({'success':True})


def temp_send_file(fileobj, filename, mimetype):
    fd,path=tempfile.mkstemp(suffix=Path(filename).suffix); os.close(fd)
    with open(path,'wb') as f:f.write(fileobj.getvalue())
    @after_this_request
    def cleanup(response):
        try:os.remove(path)
        except OSError:pass
        return response
    return send_file(path,as_attachment=True,download_name=filename,mimetype=mimetype)

@app.route('/api/backup',methods=['GET'])
@login_required
def api_backup():
    goals,history=savings_data(); expenses=get_expenses(); uid=user_id()
    fd,path=tempfile.mkstemp(suffix='.zip');os.close(fd)
    try:
        with zipfile.ZipFile(path,'w',zipfile.ZIP_DEFLATED) as z:
            eb=workbook_bytes(excel_expenses(expenses)); sb=workbook_bytes(excel_savings(goals,history)); z.writestr('Expense_Tracker_Master.xlsx',eb.getvalue());z.writestr('Savings_Tracker.xlsx',sb.getvalue())
        @after_this_request
        def cleanup(response):
            try:os.remove(path)
            except OSError:pass
            return response
        return send_file(path,as_attachment=True,download_name=f'Expense_Tracker_Backup_User_{uid}.zip',mimetype='application/zip')
    except Exception:
        try:os.remove(path)
        except OSError:pass
        raise

@app.route('/api/export/expenses',methods=['GET'])
@login_required
def api_export():
    wb=excel_expenses(get_expenses());bio=workbook_bytes(wb)
    return send_file(bio,as_attachment=True,download_name='Expense_Tracker_Master.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/settings')
@login_required
def api_settings():
    goals,history=savings_data();expenses=get_expenses()
    location='Supabase PostgreSQL (cloud)' if using_postgres() else str((DATA_DIR/'users'/str(user_id())).relative_to(BASE_DIR))
    return jsonify({'storage':{'data_location':location,'expense_file':'Generated on export','savings_file':'Generated on export','database':'PostgreSQL' if using_postgres() else 'SQLite'},'records':{'expenses':len(expenses),'savings_goals':len(goals),'savings_contributions':len(history)},'application':'Personal Expense Tracker','version':'2.0.0'})

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):return jsonify({'error':'API endpoint not found.'}),404
    return 'Page not found',404
@app.errorhandler(500)
def server_error(e):
    if request.path.startswith('/api/'):return jsonify({'error':'Internal server error.'}),500
    return 'Internal server error',500

if __name__=='__main__':
    init_database();port=int(os.environ.get('PORT',5000));app.run(host='0.0.0.0',port=port,debug=False)
